"""
Fill PLP match columns in an input XLSX from the Human and System dictionaries.

The input workbook must have these columns:
    manufacturer_part_number, manufacturer, description,
    PLP Cross Part, Confidence Score, Reason of Match, H_PLPMatch, H_MatchReason

Guard: the five "result" columns ("PLP Cross Part", "Confidence Score",
"Reason of Match", "H_PLPMatch", "H_MatchReason") must be completely empty.
If ANY cell in those columns already holds data, the file is left untouched.

For each row a lookup key is built as:
    "<manufacturer_part_number>|<manufacturer>|<description>"

Only these three columns are filled: "PLP Cross Part", "Confidence Score",
"Reason of Match".
  - First look the key up in the Human dictionary. If found, map
    H_PLPMatch -> PLP Cross Part, H_MatchReason -> Reason of Match, and copy
    its Confidence Score.
  - Otherwise look it up in the System dictionary and copy those three fields.
  - If the key is found in neither dictionary exactly, fall back to a semantic
    nearest-key search: the row key is embedded with a SentenceTransformer and
    compared (cosine similarity) against every dictionary key. The nearest key
    at or above SIMILARITY_THRESHOLD is used (Human dictionary preferred), and
    its "Reason of Match" is annotated with the matched key and similarity.
  - If nothing clears the threshold, the row is left unchanged.
"""

import json
import os
import sys

import openpyxl
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# Load configuration from a .env file (if present) into the environment.
load_dotenv()

SYSTEM_JSON = "system_dictionary.json"
HUMAN_JSON = "human_dictionary.json"


def env_bool(name, default):
    """Read a boolean from the environment (true/1/yes/on are truthy)."""
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in ("1", "true", "yes", "on")


# Configuration (overridable via .env / environment variables):
#   EMBED_MODEL           - SentenceTransformer model for the semantic fallback.
#   SIMILARITY_THRESHOLD  - minimum cosine similarity to accept a semantic match.
#   USE_SEMANTIC_SEARCH   - enable the semantic nearest-key fallback.
EMBED_MODEL = os.getenv("EMBED_MODEL", "all-MiniLM-L6-v2")
SIMILARITY_THRESHOLD = float(os.getenv("SIMILARITY_THRESHOLD", "0.80"))
USE_SEMANTIC_SEARCH = env_bool("USE_SEMANTIC_SEARCH", True)

# Columns that must be empty for the file to be processed.
RESULT_COLUMNS = [
    "PLP Cross Part",
    "Confidence Score",
    "Reason of Match",
    "H_PLPMatch",
    "H_MatchReason",
]

# The only columns this script writes into.
FILL_COLUMNS = ["PLP Cross Part", "Confidence Score", "Reason of Match"]


def clean(value):
    """Normalize a cell: treat None/whitespace as empty string."""
    if value is None:
        return ""
    return str(value).strip()


_model = None


def get_model():
    """Load the SentenceTransformer model once and reuse it."""
    global _model
    if _model is None:
        _model = SentenceTransformer(EMBED_MODEL)
    return _model


class SemanticIndex:
    """Embed a set of dictionary keys and find the nearest one by cosine."""

    def __init__(self, keys):
        self.keys = list(keys)
        self.embeddings = None
        if self.keys:
            self.embeddings = get_model().encode(
                self.keys, convert_to_numpy=True, normalize_embeddings=True
            )

    def nearest(self, query_embeddings):
        """For each query row, return (nearest_key, similarity_score).

        Returns an empty list when the index holds no keys.
        """
        if self.embeddings is None:
            return []
        sims = cosine_similarity(query_embeddings, self.embeddings)
        best_idx = sims.argmax(axis=1)
        return [(self.keys[i], float(sims[r, i])) for r, i in enumerate(best_idx)]


def load_dictionary(path):
    """Load a JSON dictionary, or return an empty dict if absent/empty."""
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            content = f.read().strip()
            if content:
                return json.loads(content)
    return {}


def header_index(ws):
    """Map header name -> 1-based column index from the first row."""
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def has_existing_data(ws, headers):
    """Return True if any RESULT_COLUMN cell (rows 2+) already holds data."""
    col_indexes = [headers[name] for name in RESULT_COLUMNS if name in headers]
    for row in ws.iter_rows(min_row=2):
        for col in col_indexes:
            if clean(row[col - 1].value):
                return True
    return False


def human_values(entry):
    """Map a Human dictionary entry to the three fill columns."""
    return {
        "PLP Cross Part": entry.get("H_PLPMatch", ""),
        "Confidence Score": entry.get("Confidence Score", ""),
        "Reason of Match": entry.get("H_MatchReason", ""),
    }


def system_values(entry):
    """Map a System dictionary entry to the three fill columns."""
    return {
        "PLP Cross Part": entry.get("PLP Cross Part", ""),
        "Confidence Score": entry.get("Confidence Score", ""),
        "Reason of Match": entry.get("Reason of Match", ""),
    }


def annotate_semantic(values, matched_key, score):
    """Note on the reason that this was a nearest-key (not exact) match."""
    note = f'[Nearest-key semantic match to "{matched_key}", cosine similarity {score:.2f}]'
    reason = clean(values.get("Reason of Match"))
    values["Reason of Match"] = f"{reason} {note}".strip()
    return values


def write_row(row, headers, values):
    """Write the three fill columns of a row from a values dict."""
    for col_name in FILL_COLUMNS:
        row[headers[col_name] - 1].value = values[col_name]


def fill_workbook(xlsx_path, system_dict, human_dict):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = header_index(ws)

    missing = [c for c in RESULT_COLUMNS if c not in headers]
    if missing:
        print(f"SKIP: missing expected column(s): {', '.join(missing)}")
        return False

    if has_existing_data(ws, headers):
        print(
            "SKIP: one or more of the result columns already contain data; "
            "input file left untouched."
        )
        return False

    key_cols = ["manufacturer_part_number", "manufacturer", "description"]
    missing_keys = [c for c in key_cols if c not in headers]
    if missing_keys:
        print(f"SKIP: missing key column(s): {', '.join(missing_keys)}")
        return False

    human_hits = 0
    system_hits = 0
    human_semantic = 0
    system_semantic = 0
    no_match = 0

    # First pass: exact-key lookups. Rows with no exact match are collected for
    # the semantic nearest-key fallback below.
    pending = []  # list of (row, key) awaiting a semantic match
    for row in ws.iter_rows(min_row=2):
        part_number = clean(row[headers["manufacturer_part_number"] - 1].value)
        manufacturer = clean(row[headers["manufacturer"] - 1].value)
        description = clean(row[headers["description"] - 1].value)

        # Skip fully blank rows.
        if not (part_number or manufacturer or description):
            continue

        key = f"{part_number}|{manufacturer}|{description}"

        if key in human_dict:
            write_row(row, headers, human_values(human_dict[key]))
            human_hits += 1
        elif key in system_dict:
            write_row(row, headers, system_values(system_dict[key]))
            system_hits += 1
        else:
            pending.append((row, key))

    # Second pass: semantic nearest-key match for rows without an exact hit.
    if pending and USE_SEMANTIC_SEARCH and (human_dict or system_dict):
        human_index = SemanticIndex(human_dict.keys())
        system_index = SemanticIndex(system_dict.keys())

        query_keys = [key for _, key in pending]
        query_embeddings = get_model().encode(
            query_keys, convert_to_numpy=True, normalize_embeddings=True
        )

        human_nearest = human_index.nearest(query_embeddings)
        system_nearest = system_index.nearest(query_embeddings)

        for i, (row, _key) in enumerate(pending):
            h_key, h_score = human_nearest[i] if human_nearest else (None, 0.0)
            s_key, s_score = system_nearest[i] if system_nearest else (None, 0.0)

            # Human dictionary is preferred when it clears the threshold.
            if h_key is not None and h_score >= SIMILARITY_THRESHOLD:
                values = annotate_semantic(
                    human_values(human_dict[h_key]), h_key, h_score
                )
                write_row(row, headers, values)
                human_semantic += 1
            elif s_key is not None and s_score >= SIMILARITY_THRESHOLD:
                values = annotate_semantic(
                    system_values(system_dict[s_key]), s_key, s_score
                )
                write_row(row, headers, values)
                system_semantic += 1
            else:
                no_match += 1  # Leave the row unchanged.
    else:
        no_match += len(pending)

    wb.save(xlsx_path)
    print(
        f"Done: {human_hits} exact + {human_semantic} semantic from Human, "
        f"{system_hits} exact + {system_semantic} semantic from System, "
        f"{no_match} unmatched -> {xlsx_path}"
    )
    return True


def main():
    if len(sys.argv) < 2:
        print("Usage: python fill_matches.py <input.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    system_dict = load_dictionary(SYSTEM_JSON)
    human_dict = load_dictionary(HUMAN_JSON)

    fill_workbook(xlsx_path, system_dict, human_dict)


if __name__ == "__main__":
    main()
